import os
import io
import csv
import re
from datetime import datetime, date
from flask import Flask, request, jsonify, render_template, g, Response
from flask_cors import CORS
from werkzeug.utils import secure_filename

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import psycopg2
    import psycopg2.extras
    PSYCOPG2_AVAILABLE = True
except ImportError:
    PSYCOPG2_AVAILABLE = False

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_ROOT_DIR = os.path.dirname(_BASE_DIR)
_STATIC_DIR = os.path.join(_ROOT_DIR, "static")
_TEMPLATE_DIR = os.path.join(_ROOT_DIR, "templates")

app = Flask(
    __name__,
    static_folder=_STATIC_DIR,
    static_url_path="/static",
    template_folder=_TEMPLATE_DIR,
)
CORS(app)

# Upload limits
MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB
MAX_IMPORT_ROWS = 2000
ALLOWED_EXTENSIONS = {"csv", "xlsx", "xls"}

# Canonical application columns for import/export
APP_COLUMNS = [
    {"key": "name", "label": "Name / Company", "required": True, "type": "text"},
    {"key": "contact_number", "label": "Contact Number", "required": True, "type": "text"},
    {"key": "source", "label": "Source", "required": True, "type": "enum"},
    {"key": "deal_value", "label": "Deal Value", "required": True, "type": "number"},
    {"key": "priority", "label": "Priority", "required": True, "type": "enum"},
    {"key": "status", "label": "Status", "required": True, "type": "enum"},
    {"key": "last_contact_date", "label": "Last Contact Date", "required": False, "type": "date"},
    {"key": "next_follow_up", "label": "Next Follow-up Date", "required": False, "type": "date"},
    {"key": "notes", "label": "Notes", "required": False, "type": "text"},
]

# Synonyms for auto column mapping (lowercase)
COLUMN_ALIASES = {
    "name": ["name", "company", "company name", "lead name", "lead", "client", "client name", "customer", "customer name"],
    "contact_number": ["contact", "contact number", "phone", "mobile", "phone number", "mobile number", "contact_number", "tel"],
    "source": ["source", "lead source", "channel", "origin"],
    "deal_value": ["deal value", "value", "amount", "deal", "deal_value", "price", "budget", "deal amount"],
    "priority": ["priority", "hotness", "urgency"],
    "status": ["status", "stage", "pipeline status"],
    "last_contact_date": ["last contact", "last contact date", "last_contact_date", "last contacted", "contacted on"],
    "next_follow_up": ["next follow up", "next follow-up", "follow up", "follow-up", "followup", "next_follow_up", "follow up date"],
    "notes": ["notes", "note", "comments", "remark", "remarks", "description"],
}

# ---------------------------------------------------------------------------
# DATABASE CONNECTION (Neon / Postgres)
#
# Your Vercel Neon integration was created with the prefix "lead_", so the
# variables it injected are named lead_DATABASE_URL, lead_PGHOST, etc.
# (not plain DATABASE_URL). We check several likely names so this works
# whether you keep the "lead_" prefix or rename them later.
# ---------------------------------------------------------------------------
DATABASE_URL = (
    os.environ.get("lead_DATABASE_URL")
    or os.environ.get("lead_POSTGRES_URL")
    or os.environ.get("POSTGRES_URL")
    or os.environ.get("DATABASE_URL")
)

if not DATABASE_URL:
    env_file = os.path.join(_ROOT_DIR, ".env")
    if os.path.exists(env_file):
        with open(env_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" in line:
                    k, v = line.split("=", 1)
                    k, v = k.strip(), v.strip().strip("'\"")
                    os.environ[k] = v
        DATABASE_URL = (
            os.environ.get("lead_DATABASE_URL")
            or os.environ.get("lead_POSTGRES_URL")
            or os.environ.get("POSTGRES_URL")
            or os.environ.get("DATABASE_URL")
        )

if not DATABASE_URL:
    raise RuntimeError(
        "No Postgres connection string found. Set the 'lead_DATABASE_URL' "
        "environment variable in Vercel (Project -> Settings -> Environment "
        "Variables) to your Neon connection string."
    )


def get_db():
    db = getattr(g, "_database", None)
    if db is None:
        if not PSYCOPG2_AVAILABLE:
            raise RuntimeError("psycopg2-binary is not installed. Add it to requirements.txt")
        db = g._database = psycopg2.connect(
            DATABASE_URL,
            cursor_factory=psycopg2.extras.RealDictCursor,
        )
    return db


@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, "_database", None)
    if db is not None:
        db.close()


_db_initialized = False


def init_db(force=False):
    """Create tables, indexes, and seed demo data if empty. Safe and memoized per process."""
    global _db_initialized
    if _db_initialized and not force:
        return
    db = get_db()
    with db.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS leads (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                contact_number TEXT NOT NULL,
                source TEXT NOT NULL,
                deal_value NUMERIC NOT NULL DEFAULT 0,
                priority TEXT NOT NULL CHECK (priority IN ('Hot', 'Warm', 'Cold')),
                status TEXT NOT NULL CHECK (status IN ('New', 'Contacted', 'Converted', 'Lost')),
                last_contact_date DATE,
                next_follow_up DATE,
                notes TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_leads_status ON leads (status);
            CREATE INDEX IF NOT EXISTS idx_leads_priority ON leads (priority);
            CREATE INDEX IF NOT EXISTS idx_leads_next_follow_up ON leads (next_follow_up);
            CREATE INDEX IF NOT EXISTS idx_leads_created_at ON leads (created_at);
        """)
        db.commit()

        cur.execute("SELECT COUNT(*) AS c FROM leads")
        count = cur.fetchone()["c"]

        if count == 0:
            now = datetime.utcnow().isoformat() + "Z"
            today = date.today()
            demo_leads = [
                (
                    "Sigma Electric", "+91 98200 11223", "Referral", 850000, "Hot", "Contacted",
                    (today.replace(day=max(1, today.day - 2))).isoformat(),
                    (today.replace(day=max(1, today.day - 4))).isoformat(),
                    "Called customer on Monday. Interested in enterprise package. Requested pricing proposal. Follow up ASAP.",
                    now, now,
                ),
                (
                    "B-Infra Hyderabad", "+91 90000 44556", "Outbound", 400000, "Warm", "New",
                    None,
                    (today.replace(day=min(28, today.day + 2))).isoformat(),
                    "Initial outreach via LinkedIn. Decision maker is the Project Director.",
                    now, now,
                ),
                (
                    "Aether Tech Solutions", "+91 98765 43210", "Website", 1250000, "Hot", "Contacted",
                    (today.replace(day=max(1, today.day - 1))).isoformat(),
                    (today.replace(day=min(28, today.day + 1))).isoformat(),
                    "Demo completed. Waiting for internal budget approval. High intent.",
                    now, now,
                ),
                (
                    "Nova Retail Pvt Ltd", "+91 91234 56789", "Social Media", 275000, "Cold", "New",
                    None,
                    (today.replace(day=min(28, today.day + 5))).isoformat(),
                    "Inbound inquiry from Instagram ad. Needs more nurturing.",
                    now, now,
                ),
                (
                    "Horizon Logistics", "+91 99887 76655", "Existing Customer", 620000, "Warm", "Contacted",
                    (today.replace(day=max(1, today.day - 5))).isoformat(),
                    (today.replace(day=max(1, today.day - 1))).isoformat(),
                    "Upsell opportunity for warehouse management module. Spoke with ops head.",
                    now, now,
                ),
                (
                    "Pixel Design Studio", "+91 97654 32109", "Referral", 180000, "Warm", "Converted",
                    (today.replace(day=max(1, today.day - 10))).isoformat(),
                    None,
                    "Deal closed. Implementation starts next month. Happy customer.",
                    now, now,
                ),
                (
                    "GreenField Agro", "+91 96543 21098", "Advertisement", 950000, "Hot", "New",
                    None,
                    (today.replace(day=min(28, today.day + 3))).isoformat(),
                    "Responded to Google Ads campaign. Looking for ERP for 3 locations.",
                    now, now,
                ),
                (
                    "Metro Construction Co", "+91 95432 10987", "Outbound", 2100000, "Hot", "Contacted",
                    (today.replace(day=max(1, today.day - 3))).isoformat(),
                    (today.replace(day=max(1, today.day - 2))).isoformat(),
                    "Multiple stakeholders. Need to send revised proposal with volume discount.",
                    now, now,
                ),
                (
                    "Lumina Healthcare", "+91 94321 09876", "Website", 540000, "Warm", "Lost",
                    (today.replace(day=max(1, today.day - 20))).isoformat(),
                    None,
                    "Chose competitor due to tighter timeline. Keep warm for future.",
                    now, now,
                ),
                (
                    "SwiftPay Fintech", "+91 93210 98765", "Other", 780000, "Cold", "New",
                    None,
                    (today.replace(day=min(28, today.day + 7))).isoformat(),
                    "Came through industry event. Early stage discussion.",
                    now, now,
                ),
            ]
            cur.executemany(
                """
                INSERT INTO leads (
                    name, contact_number, source, deal_value, priority, status,
                    last_contact_date, next_follow_up, notes, created_at, updated_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                demo_leads,
            )
            db.commit()
    _db_initialized = True


def row_to_dict(row):
    if row is None:
        return None
    return dict(row)


def validate_lead_data(data, is_update=False):
    errors = []
    name = (data.get("name") or "").strip()
    if not name:
        errors.append("Name / Company Name is required")
    elif len(name) > 200:
        errors.append("Name must be under 200 characters")

    contact = (data.get("contact_number") or "").strip()
    if not contact:
        errors.append("Contact Number is required")
    elif len(contact) < 8 or len(contact) > 20:
        errors.append("Contact Number looks invalid")

    source = (data.get("source") or "").strip()
    allowed_sources = [
        "Referral", "Website", "Outbound", "Social Media",
        "Advertisement", "Existing Customer", "Other"
    ]
    if not source:
        errors.append("Source is required")
    elif source not in allowed_sources:
        errors.append("Invalid source value")

    try:
        deal_value = float(data.get("deal_value", 0))
        if deal_value < 0:
            errors.append("Deal Value cannot be negative")
        if deal_value > 1e12:
            errors.append("Deal Value is unrealistically large")
    except (TypeError, ValueError):
        errors.append("Deal Value must be a valid number")
        deal_value = 0

    priority = (data.get("priority") or "").strip()
    if priority not in ("Hot", "Warm", "Cold"):
        errors.append("Priority must be Hot, Warm or Cold")

    status = (data.get("status") or "").strip()
    if status not in ("New", "Contacted", "Converted", "Lost"):
        errors.append("Status must be New, Contacted, Converted or Lost")

    last_contact = data.get("last_contact_date")
    if last_contact:
        last_contact = str(last_contact).strip()
        if last_contact:
            try:
                datetime.strptime(last_contact, "%Y-%m-%d")
            except ValueError:
                errors.append("Last Contact Date must be YYYY-MM-DD")
        else:
            last_contact = None
    else:
        last_contact = None

    next_follow = data.get("next_follow_up")
    if next_follow:
        next_follow = str(next_follow).strip()
        if next_follow:
            try:
                datetime.strptime(next_follow, "%Y-%m-%d")
            except ValueError:
                errors.append("Next Follow-up Date must be YYYY-MM-DD")
        else:
            next_follow = None
    else:
        next_follow = None

    notes = (data.get("notes") or "").strip()
    if len(notes) > 5000:
        errors.append("Notes must be under 5000 characters")

    if errors:
        return None, errors

    return {
        "name": name,
        "contact_number": contact,
        "source": source,
        "deal_value": deal_value,
        "priority": priority,
        "status": status,
        "last_contact_date": last_contact,
        "next_follow_up": next_follow,
        "notes": notes,
    }, None


def compute_days_since_contact(last_contact_date):
    if not last_contact_date:
        return None
    try:
        if hasattr(last_contact_date, "isoformat") and not isinstance(last_contact_date, str):
            lcd = last_contact_date
        else:
            lcd = datetime.strptime(str(last_contact_date), "%Y-%m-%d").date()
        return (date.today() - lcd).days
    except Exception:
        return None


def is_overdue(next_follow_up, status):
    if not next_follow_up:
        return False
    if status in ("Converted", "Lost"):
        return False
    try:
        if hasattr(next_follow_up, "isoformat") and not isinstance(next_follow_up, str):
            nfd = next_follow_up
        else:
            nfd = datetime.strptime(str(next_follow_up), "%Y-%m-%d").date()
        return nfd < date.today()
    except Exception:
        return False


def _date_str(v):
    """Postgres returns DATE columns as datetime.date objects; normalize to 'YYYY-MM-DD' string for the frontend."""
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        return v.isoformat()
    return str(v)


def enrich_lead(lead):
    """Add computed fields for the frontend and normalize date types."""
    if lead is None:
        return None
    lead = dict(lead)
    lead["last_contact_date"] = _date_str(lead.get("last_contact_date"))
    lead["next_follow_up"] = _date_str(lead.get("next_follow_up"))
    days = compute_days_since_contact(lead.get("last_contact_date"))
    lead["days_since_contact"] = days
    lead["is_overdue"] = is_overdue(lead.get("next_follow_up"), lead.get("status"))
    if lead.get("deal_value") is not None:
        lead["deal_value"] = float(lead["deal_value"])
    return lead


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/leads", methods=["GET"])
def get_leads():
    init_db()
    db = get_db()
    status = request.args.get("status")
    priority = request.args.get("priority")
    search = (request.args.get("search") or "").strip()
    sort = request.args.get("sort", "attention")

    query = "SELECT * FROM leads WHERE 1=1"
    params = []
    if status and status != "All":
        query += " AND status = %s"
        params.append(status)
    if priority and priority != "All":
        query += " AND priority = %s"
        params.append(priority)
    if search:
        query += " AND (name ILIKE %s OR contact_number ILIKE %s OR source ILIKE %s OR notes ILIKE %s)"
        like = f"%{search}%"
        params.extend([like, like, like, like])

    if sort == "newest":
        query += " ORDER BY created_at DESC"
    elif sort == "oldest":
        query += " ORDER BY created_at ASC"
    elif sort == "value_high":
        query += " ORDER BY deal_value DESC"
    elif sort == "value_low":
        query += " ORDER BY deal_value ASC"
    elif sort == "followup":
        query += " ORDER BY CASE WHEN next_follow_up IS NULL THEN 1 ELSE 0 END, next_follow_up ASC"
    elif sort == "priority":
        query += " ORDER BY CASE priority WHEN 'Hot' THEN 1 WHEN 'Warm' THEN 2 ELSE 3 END"
    else:
        query += """
            ORDER BY
                CASE
                    WHEN next_follow_up IS NOT NULL
                         AND next_follow_up < CURRENT_DATE
                         AND status NOT IN ('Converted', 'Lost')
                    THEN 0
                    ELSE 1
                END,
                CASE WHEN next_follow_up IS NULL THEN 1 ELSE 0 END,
                next_follow_up ASC,
                created_at DESC
        """

    with db.cursor() as cur:
        cur.execute(query, params)
        rows = cur.fetchall()
    leads = [enrich_lead(row) for row in rows]
    return jsonify({"success": True, "data": leads})


@app.route("/api/leads/<int:lead_id>", methods=["GET"])
def get_lead(lead_id):
    init_db()
    db = get_db()
    with db.cursor() as cur:
        cur.execute("SELECT * FROM leads WHERE id = %s", (lead_id,))
        row = cur.fetchone()
    if not row:
        return jsonify({"success": False, "error": "Lead not found"}), 404
    return jsonify({"success": True, "data": enrich_lead(row)})


@app.route("/api/leads", methods=["POST"])
def create_lead():
    init_db()
    data = request.get_json(silent=True) or {}
    validated, errors = validate_lead_data(data)
    if errors:
        return jsonify({"success": False, "errors": errors}), 400

    now = datetime.utcnow().isoformat() + "Z"
    db = get_db()
    with db.cursor() as cur:
        cur.execute(
            """
            INSERT INTO leads (
                name, contact_number, source, deal_value, priority, status,
                last_contact_date, next_follow_up, notes, created_at, updated_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
            """,
            (
                validated["name"], validated["contact_number"], validated["source"],
                validated["deal_value"], validated["priority"], validated["status"],
                validated["last_contact_date"], validated["next_follow_up"],
                validated["notes"], now, now,
            ),
        )
        lead_id = cur.fetchone()["id"]
        db.commit()
        cur.execute("SELECT * FROM leads WHERE id = %s", (lead_id,))
        row = cur.fetchone()
    return jsonify({"success": True, "data": enrich_lead(row)}), 201


@app.route("/api/leads/<int:lead_id>", methods=["PUT"])
def update_lead(lead_id):
    init_db()
    db = get_db()
    with db.cursor() as cur:
        cur.execute("SELECT * FROM leads WHERE id = %s", (lead_id,))
        existing = cur.fetchone()
        if not existing:
            return jsonify({"success": False, "error": "Lead not found"}), 404

        data = request.get_json(silent=True) or {}
        validated, errors = validate_lead_data(data, is_update=True)
        if errors:
            return jsonify({"success": False, "errors": errors}), 400

        now = datetime.utcnow().isoformat() + "Z"
        cur.execute(
            """
            UPDATE leads SET
                name = %s, contact_number = %s, source = %s, deal_value = %s,
                priority = %s, status = %s, last_contact_date = %s, next_follow_up = %s,
                notes = %s, updated_at = %s
            WHERE id = %s
            """,
            (
                validated["name"], validated["contact_number"], validated["source"],
                validated["deal_value"], validated["priority"], validated["status"],
                validated["last_contact_date"], validated["next_follow_up"],
                validated["notes"], now, lead_id,
            ),
        )
        db.commit()
        cur.execute("SELECT * FROM leads WHERE id = %s", (lead_id,))
        row = cur.fetchone()
    return jsonify({"success": True, "data": enrich_lead(row)})


@app.route("/api/leads/<int:lead_id>", methods=["DELETE"])
def delete_lead(lead_id):
    init_db()
    db = get_db()
    with db.cursor() as cur:
        cur.execute("SELECT * FROM leads WHERE id = %s", (lead_id,))
        existing = cur.fetchone()
        if not existing:
            return jsonify({"success": False, "error": "Lead not found"}), 404
        cur.execute("DELETE FROM leads WHERE id = %s", (lead_id,))
        db.commit()
    return jsonify({"success": True, "message": "Lead deleted"})


@app.route("/api/stats", methods=["GET"])
def get_stats():
    init_db()
    db = get_db()
    with db.cursor() as cur:
        cur.execute(
            """
            SELECT
                COUNT(*) AS total,
                COUNT(*) FILTER (
                    WHERE next_follow_up IS NOT NULL
                      AND next_follow_up < CURRENT_DATE
                      AND status NOT IN ('Converted', 'Lost')
                ) AS overdue,
                COALESCE(SUM(deal_value) FILTER (WHERE status != 'Lost'), 0) AS pipeline,
                COUNT(*) FILTER (WHERE status = 'New') AS count_new,
                COUNT(*) FILTER (WHERE status = 'Contacted') AS count_contacted,
                COUNT(*) FILTER (WHERE status = 'Converted') AS count_converted,
                COUNT(*) FILTER (WHERE status = 'Lost') AS count_lost,
                COUNT(*) FILTER (WHERE priority = 'Hot') AS count_hot,
                COUNT(*) FILTER (WHERE priority = 'Warm') AS count_warm,
                COUNT(*) FILTER (WHERE priority = 'Cold') AS count_cold
            FROM leads
            """
        )
        row = cur.fetchone()
        total = row["total"] or 0
        overdue = row["overdue"] or 0
        pipeline = float(row["pipeline"] or 0)
        status_counts = {
            "New": row["count_new"] or 0,
            "Contacted": row["count_contacted"] or 0,
            "Converted": row["count_converted"] or 0,
            "Lost": row["count_lost"] or 0,
        }
        priority_counts = {
            "Hot": row["count_hot"] or 0,
            "Warm": row["count_warm"] or 0,
            "Cold": row["count_cold"] or 0,
        }

    return jsonify({
        "success": True,
        "data": {
            "total_leads": total,
            "overdue_followups": overdue,
            "pipeline_value": pipeline,
            "status_counts": status_counts,
            "priority_counts": priority_counts,
        }
    })


# ========== Import / Export helpers ==========

def _normalize_header(h):
    if h is None:
        return ""
    return re.sub(r"\s+", " ", str(h).strip().lower())


def auto_map_columns(headers):
    mapping = {}
    used_indices = set()
    normalized = [_normalize_header(h) for h in headers]
    for col in APP_COLUMNS:
        key = col["key"]
        aliases = COLUMN_ALIASES.get(key, [key])
        for i, nh in enumerate(normalized):
            if i in used_indices:
                continue
            if nh in aliases or nh.replace("_", " ") in aliases:
                mapping[key] = i
                used_indices.add(i)
                break
    return mapping


def parse_date_value(val):
    if val is None or str(val).strip() == "":
        return None
    s = str(val).strip()
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(val).date().isoformat()
        except Exception:
            pass
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(s[:20], fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.replace("Z", "").split("T")[0]).date().isoformat()
    except Exception:
        return None


def parse_deal_value(val):
    if val is None or str(val).strip() == "":
        return None
    s = str(val).strip().replace(",", "").replace("₹", "").replace("Rs.", "").replace("Rs", "")
    s = s.replace("L", "").replace("l", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def read_uploaded_rows(file_storage):
    filename = secure_filename(file_storage.filename or "")
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in ALLOWED_EXTENSIONS:
        return None, None, "Only .csv, .xlsx, .xls files are allowed"
    raw = file_storage.read()
    if len(raw) > MAX_UPLOAD_BYTES:
        return None, None, f"File too large. Max {MAX_UPLOAD_BYTES // (1024*1024)} MB"
    if len(raw) == 0:
        return None, None, "File is empty"

    if ext == "csv":
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = raw.decode("latin-1")
            except Exception:
                return None, None, "Could not decode CSV file"
        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)
        if not all_rows:
            return None, None, "CSV has no rows"
        headers = [str(h).strip() for h in all_rows[0]]
        data_rows = all_rows[1:MAX_IMPORT_ROWS + 1]
        return headers, data_rows, None

    if not OPENPYXL_AVAILABLE:
        return None, None, "Excel support requires openpyxl. Run: pip install openpyxl"
    if ext == "xls":
        return None, None, "Legacy .xls is not supported. Please save as .xlsx or .csv"
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        all_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > MAX_IMPORT_ROWS:
                break
            all_rows.append(list(row))
        wb.close()
        if not all_rows:
            return None, None, "Excel sheet is empty"
        headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]
        data_rows = all_rows[1:]
        return headers, data_rows, None
    except Exception as e:
        return None, None, f"Failed to read Excel file: {str(e)}"


def row_to_lead_dict(row, mapping):
    def cell(key):
        idx = mapping.get(key)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return None if v is None else v

    name = cell("name")
    contact = cell("contact_number")
    source = cell("source")
    deal_raw = cell("deal_value")
    priority = cell("priority")
    status = cell("status")
    last_c = cell("last_contact_date")
    next_f = cell("next_follow_up")
    notes = cell("notes")

    data = {
        "name": str(name).strip() if name is not None else "",
        "contact_number": str(contact).strip() if contact is not None else "",
        "source": str(source).strip() if source is not None else "",
        "deal_value": parse_deal_value(deal_raw),
        "priority": str(priority).strip().title() if priority is not None else "",
        "status": str(status).strip().title() if status is not None else "",
        "last_contact_date": parse_date_value(last_c),
        "next_follow_up": parse_date_value(next_f),
        "notes": str(notes).strip() if notes is not None else "",
    }

    if data["priority"] not in ("Hot", "Warm", "Cold"):
        p = data["priority"].lower()
        if p in ("high", "hot"):
            data["priority"] = "Hot"
        elif p in ("medium", "warm", "mid"):
            data["priority"] = "Warm"
        elif p in ("low", "cold"):
            data["priority"] = "Cold"

    if data["status"] not in ("New", "Contacted", "Converted", "Lost"):
        s = data["status"].lower()
        if s in ("new", "fresh"):
            data["status"] = "New"
        elif s in ("contacted", "in progress", "follow up"):
            data["status"] = "Contacted"
        elif s in ("converted", "won", "closed"):
            data["status"] = "Converted"
        elif s in ("lost", "dead"):
            data["status"] = "Lost"

    allowed_sources = [
        "Referral", "Website", "Outbound", "Social Media",
        "Advertisement", "Existing Customer", "Other"
    ]
    if data["source"] and data["source"] not in allowed_sources:
        found = None
        for a in allowed_sources:
            if a.lower() == data["source"].lower():
                found = a
                break
        data["source"] = found or "Other"

    return data


def validate_import_row(data, existing_keys):
    errors = []
    if not data.get("name"):
        errors.append("Name is required")
    if not data.get("contact_number"):
        errors.append("Contact number is required")
    if data.get("deal_value") is None:
        errors.append("Deal value must be a number")
    elif data["deal_value"] < 0:
        errors.append("Deal value cannot be negative")
    if data.get("priority") not in ("Hot", "Warm", "Cold"):
        errors.append("Priority must be Hot, Warm or Cold")
    if data.get("status") not in ("New", "Contacted", "Converted", "Lost"):
        errors.append("Status must be New, Contacted, Converted or Lost")
    if not data.get("source"):
        errors.append("Source is required")
    key = (data.get("name", "").lower(), data.get("contact_number", "").lower())
    if key in existing_keys:
        errors.append("Duplicate (same name + contact already exists or in file)")
    return errors


@app.route("/api/import/preview", methods=["POST"])
def import_preview():
    init_db()
    if "file" not in request.files:
        return jsonify({"success": False, "error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"success": False, "error": "No file selected"}), 400

    headers, data_rows, err = read_uploaded_rows(f)
    if err:
        return jsonify({"success": False, "error": err}), 400

    manual_map = None
    if request.form.get("mapping"):
        try:
            import json
            manual_map = json.loads(request.form.get("mapping"))
        except Exception:
            manual_map = None

    if manual_map:
        header_to_idx = {_normalize_header(h): i for i, h in enumerate(headers)}
        mapping = {}
        for key, header_name in manual_map.items():
            if not header_name:
                continue
            idx = header_to_idx.get(_normalize_header(header_name))
            if idx is not None:
                mapping[key] = idx
    else:
        mapping = auto_map_columns(headers)

    mapped_keys = set(mapping.keys())
    all_keys = {c["key"] for c in APP_COLUMNS}
    required_keys = {c["key"] for c in APP_COLUMNS if c["required"]}
    missing_required = sorted(required_keys - mapped_keys)
    missing_optional = sorted((all_keys - required_keys) - mapped_keys)
    used_indices = set(mapping.values())
    extra_columns = [headers[i] for i in range(len(headers)) if i not in used_indices and headers[i]]

    db = get_db()
    with db.cursor() as cur:
        cur.execute("SELECT name, contact_number FROM leads")
        existing = cur.fetchall()
    existing_keys = {(r["name"].lower(), r["contact_number"].lower()) for r in existing}

    preview_rows = []
    valid_count = 0
    invalid_count = 0
    duplicate_count = 0
    file_keys = set()

    for i, row in enumerate(data_rows):
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        data = row_to_lead_dict(row, mapping)
        key = (data.get("name", "").lower(), data.get("contact_number", "").lower())
        errors = validate_import_row(data, existing_keys | file_keys)
        is_dup = any("Duplicate" in e for e in errors)
        if is_dup:
            duplicate_count += 1
        if errors:
            invalid_count += 1
        else:
            valid_count += 1
            file_keys.add(key)
        preview_rows.append({
            "row_number": i + 2,
            "data": data,
            "errors": errors,
            "valid": len(errors) == 0,
        })

    valid_payload = [r["data"] for r in preview_rows if r["valid"]]
    return jsonify({
        "success": True,
        "data": {
            "headers": headers,
            "auto_mapping": {k: headers[v] for k, v in mapping.items()},
            "app_columns": APP_COLUMNS,
            "missing_required": missing_required,
            "missing_optional": missing_optional,
            "extra_columns": extra_columns,
            "total_rows": len(preview_rows),
            "valid_rows": valid_count,
            "invalid_rows": invalid_count,
            "duplicate_rows": duplicate_count,
            "preview": preview_rows[:100],
            "valid_payload": valid_payload,
            "mapping": mapping,
        }
    })


@app.route("/api/import/confirm", methods=["POST"])
def import_confirm():
    init_db()
    body = request.get_json(silent=True) or {}
    rows = body.get("rows") or []
    if not isinstance(rows, list):
        return jsonify({"success": False, "error": "Invalid payload"}), 400
    if len(rows) > MAX_IMPORT_ROWS:
        return jsonify({"success": False, "error": f"Max {MAX_IMPORT_ROWS} rows per import"}), 400

    db = get_db()
    with db.cursor() as cur:
        cur.execute("SELECT name, contact_number FROM leads")
        existing = cur.fetchall()
    existing_keys = {(r["name"].lower(), r["contact_number"].lower()) for r in existing}

    inserted = 0
    skipped = 0
    errors_out = []
    now = datetime.utcnow().isoformat() + "Z"

    try:
        with db.cursor() as cur:
            for i, raw in enumerate(rows):
                validated, errs = validate_lead_data(raw)
                if errs:
                    skipped += 1
                    errors_out.append({"row": i + 1, "errors": errs})
                    continue
                key = (validated["name"].lower(), validated["contact_number"].lower())
                if key in existing_keys:
                    skipped += 1
                    errors_out.append({"row": i + 1, "errors": ["Duplicate"]})
                    continue
                cur.execute(
                    """
                    INSERT INTO leads (
                        name, contact_number, source, deal_value, priority, status,
                        last_contact_date, next_follow_up, notes, created_at, updated_at
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        validated["name"], validated["contact_number"], validated["source"],
                        validated["deal_value"], validated["priority"], validated["status"],
                        validated["last_contact_date"], validated["next_follow_up"],
                        validated["notes"], now, now,
                    ),
                )
                existing_keys.add(key)
                inserted += 1
        db.commit()
    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "error": "Import failed: " + str(e)}), 500

    return jsonify({
        "success": True,
        "data": {
            "inserted": inserted,
            "skipped": skipped,
            "errors": errors_out[:50],
        }
    })


@app.route("/api/export", methods=["GET"])
def export_leads():
    init_db()
    fmt = (request.args.get("format") or "csv").lower()
    if fmt not in ("csv", "xlsx"):
        return jsonify({"success": False, "error": "format must be csv or xlsx"}), 400

    status = request.args.get("status")
    priority = request.args.get("priority")
    search = (request.args.get("search") or "").strip()
    sort = request.args.get("sort", "attention")

    query = "SELECT * FROM leads WHERE 1=1"
    params = []
    if status and status != "All":
        query += " AND status = %s"
        params.append(status)
    if priority and priority != "All":
        query += " AND priority = %s"
        params.append(priority)
    if search:
        query += " AND (name ILIKE %s OR contact_number ILIKE %s OR source ILIKE %s OR notes ILIKE %s)"
        like = f"%{search}%"
        params.extend([like, like, like, like])

    if sort == "newest":
        query += " ORDER BY created_at DESC"
    elif sort == "oldest":
        query += " ORDER BY created_at ASC"
    elif sort == "value_high":
        query += " ORDER BY deal_value DESC"
    elif sort == "value_low":
        query += " ORDER BY deal_value ASC"
    elif sort == "followup":
        query += " ORDER BY CASE WHEN next_follow_up IS NULL THEN 1 ELSE 0 END, next_follow_up ASC"
    elif sort == "priority":
        query += " ORDER BY CASE priority WHEN 'Hot' THEN 1 WHEN 'Warm' THEN 2 ELSE 3 END"
    else:
        query += """
            ORDER BY
                CASE
                    WHEN next_follow_up IS NOT NULL
                         AND next_follow_up < CURRENT_DATE
                         AND status NOT IN ('Converted', 'Lost')
                    THEN 0 ELSE 1 END,
                CASE WHEN next_follow_up IS NULL THEN 1 ELSE 0 END,
                next_follow_up ASC, created_at DESC
        """

    db = get_db()
    with db.cursor() as cur:
        cur.execute(query, params)
        rows = cur.fetchall()
    leads = [enrich_lead(r) for r in rows]

    today = date.today().isoformat()
    filter_tag = "filtered_" if (status and status != "All") or (priority and priority != "All") or search else ""
    filename_base = f"leads_{filter_tag}{today}"
    export_headers = [
        "Name / Company", "Contact Number", "Source", "Deal Value",
        "Priority", "Status", "Last Contact Date", "Next Follow-up",
        "Days Since Contact", "Notes", "Created At", "Updated At"
    ]

    def lead_row(l):
        return [
            l.get("name") or "",
            l.get("contact_number") or "",
            l.get("source") or "",
            l.get("deal_value") if l.get("deal_value") is not None else "",
            l.get("priority") or "",
            l.get("status") or "",
            l.get("last_contact_date") or "",
            l.get("next_follow_up") or "",
            l.get("days_since_contact") if l.get("days_since_contact") is not None else "",
            l.get("notes") or "",
            (l.get("created_at") or "")[:19],
            (l.get("updated_at") or "")[:19],
        ]

    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(export_headers)
        for l in leads:
            writer.writerow(lead_row(l))
        output = buf.getvalue()
        return Response(
            output,
            mimetype="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="{filename_base}.csv"',
                "Content-Type": "text/csv; charset=utf-8",
            },
        )

    if not OPENPYXL_AVAILABLE:
        return jsonify({"success": False, "error": "openpyxl required for Excel export"}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    thin = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    for col, h in enumerate(export_headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
    for r_idx, l in enumerate(leads, 2):
        for c_idx, val in enumerate(lead_row(l), 1):
            cell = ws.cell(r_idx, c_idx, val)
            cell.border = thin
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename_base}.xlsx"',
        },
    )


# For local development
if __name__ == "__main__":
    with app.app_context():
        init_db()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
